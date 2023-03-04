# -*- coding: utf-8 -*-
# @Author   : Fricky
# @Time     : 2022-09-07 10:44
import argparse
import time
import openpyxl
import requests
from jsonpath import jsonpath
import os

from pyfiglet import Figlet

url = "https://api.threatbook.cn/v3/scene/ip_reputation"


def judge_code(code):
    if code == 0:
        return
    elif code == -1:
        print("请检查APIKEY！")
        exit()
    elif code == -4:
        print("超出请求限制！请登录微步官网查看API余额\r\n"
              "https://x.threatbook.com/v5/myApi")
        exit()
    else:
        print("接口出错了！")
        exit()


def req(input_file, output_file):
    # ips = [i for i in open('ip.txt', 'r')]

    # 读取apikey
    key_file = open('config', 'r')
    api_key = key_file.read().strip().split('\n')

    # 判断输出文件是否存在
    if os.path.exists(f'{output_file}'):
        print(f"文件{output_file}已存在！")
        dele = input("是否删除文件 Y/N: ")

        if dele.strip().upper() == "Y":
            os.remove(output_file)
        else:
            print("输入有误，请重试！")
            exit()

    # 创建execl文件
    wb = openpyxl.Workbook()
    wb.save(output_file)
    read_xlsx = openpyxl.load_workbook(output_file)  # 读取文件
    readSheet = read_xlsx["Sheet"]  # 使用Sheet工作表
    # readSheet.delete_rows(readSheet.min_row, readSheet.max_row)  # 清空文件内容
    readSheet.append(['IP地址', '情报可信度评分', '应用场景', '威胁类型', 'IP归属地'])

    # 循环请求api
    num = 0
    for ip in open(f'./{input_file}', 'r'):
        ip = ip.strip()
        # if num % 60 == 0:
        #     print("Api超频等待5秒...")
        #     time.sleep(15)
        num += 1
        query = {
            "apikey": api_key,  # apikey
            "resource": ip,
            "lang": "zh"
        }
        try:
            response = requests.request("GET", url, params=query)
            result = response.json()

            # print(type(jsonpath(result, "$..response_code")[0]))
            judge_code(jsonpath(result, "$..response_code")[0])  # 判断api状态
            save_excel(num, ip, result, output_file, read_xlsx, readSheet)  # 保存结果
            # print(response.status_code)
            # print(result)
        except Exception as e:
            print(e)


def save_excel(num, ip, result, output_file, read_xlsx, readSheet):
    a2 = jsonpath(result, "$..confidence_level")[0]  # 可信度评分
    a3 = ",".join(jsonpath(result, "$..scene"))  # 应用场景
    a4 = ",".join(jsonpath(result, "$..judgments")[0])  # 威胁类型
    a5 = "-".join(jsonpath(result, "$..location.*")[0:3])  # IP归属地
    # a6 = ",".join(jsonpath(result, "$..tag"))
    # print(a2)
    # print(a3)
    # print(a4)
    # print(a5)
    print(f"[{num}]", '{:<15}'.format(ip.strip()), a2.strip(), a3.strip(), a4.strip(), a5.strip())

    line = [ip, a2, a3, a4, a5]
    readSheet.append(line)
    read_xlsx.save(output_file)


if __name__ == '__main__':
    banner = Figlet(width=2000)
    print(banner.renderText('IP Analysis'))

    parser = argparse.ArgumentParser(description='Simple Port Scan By Fricky. V 2.0 - 22.11.28')
    # parser.add_argument('-ip', dest='input_ip', help='Specify the IP to analyze')
    parser.add_argument('-if', dest='ip_file', help='Specifies the text file to save the IP')
    # parser.add_argument('-t', dest='thread', default=1, help='Specifying the number of threads')
    parser.add_argument('-of', dest='output_file', help='Specifies the output CSV file name')
    args = parser.parse_args()
    if args.ip_file and args.output_file:
        req(args.ip_file, args.output_file)
    else:
        print("请重新输入！")
